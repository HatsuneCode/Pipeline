# -*- coding: utf-8 -*-
"""
analyze_unified.py
统一 PubMed 文献相关性分析框架
支持领域：COPD / 神经退行性疾病 / RNA剪接 / 虚拟细胞

用法：
    python analyze_unified.py              # 运行全部领域
    python analyze_unified.py copd         # 仅运行 COPD
    python analyze_unified.py neuro rna    # 运行多个领域
    python analyze_unified.py vcell        # 仅运行虚拟细胞

可用领域标识：copd / neuro / rna / vcell
"""
import sys
import re
import openpyxl
from collections import defaultdict
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════
# 输入文件路径（统一配置）
# ══════════════════════════════════════════════════════════════════════
SRC_PATH = 'E:/Project/BioInfo/temp/sets.xlsx'


# ══════════════════════════════════════════════════════════════════════
# 共享工具函数（所有领域通用，逻辑与原三脚本完全一致）
# ══════════════════════════════════════════════════════════════════════
def has_context(text_l, context_patterns):
    return any(re.search(p, text_l, re.IGNORECASE) for p in context_patterns)


def find_hits(text_l, patterns):
    """返回 (名称, 匹配词, 上下文片段, 命中次数) 四元组列表"""
    hits = []
    for pat, name in patterns:
        matches = list(re.finditer(pat, text_l, re.IGNORECASE))
        if matches:
            m = matches[0]
            s = max(0, m.start() - 35)
            e = min(len(text_l), m.end() + 35)
            snippet = '...' + text_l[s:e].strip().replace('\n', ' ') + '...'
            hits.append((name, m.group(), snippet, len(matches)))
    return hits


def best_snippet(hits):
    if not hits:
        return ''
    return max(hits, key=lambda h: h[3])[2]


def detect_false_positives(text_l, fp_patterns):
    return [(name, pat) for pat, name in fp_patterns if re.search(pat, text_l, re.IGNORECASE)]


def classify_category(text_l, is_related, category_patterns, fallback_label):
    if not is_related:
        return '—'
    counts = {cat: sum(1 for p in pats if re.search(p, text_l, re.IGNORECASE))
              for cat, pats in category_patterns.items()}
    matched = [cat for cat, cnt in counts.items() if cnt >= 2]
    if not matched:
        best_cat = max(counts, key=counts.get)
        matched = [best_cat] if counts[best_cat] >= 1 else []
    return ' / '.join(matched) if matched else fallback_label


def detect_species(text_l, species_patterns):
    found = []
    for sp, pats in species_patterns.items():
        for p in pats:
            if re.search(p, text_l, re.IGNORECASE):
                found.append(sp)
                break
    return '、'.join(found) if found else '未明确'


def calc_score(t1, t2, t3, title_l, fp_count, is_related):
    if not is_related:
        return 0
    score = 0
    if t1:
        score += 4
        if len(set(h[0] for h in t1)) > 1:
            score += 2
        if sum(h[3] for h in t1) >= 5:
            score += 1
        if any(h[1] in title_l for h in t1):
            score += 2
    elif t2:
        score += 3
        if len(set(h[0] for h in t2)) > 1:
            score += 1
        if any(h[1] in title_l for h in t2):
            score += 1
    elif t3:
        score += 2
    if not t1:
        score = max(0, score - fp_count)
    return min(10, score)


def score_to_level(score, is_related):
    if not is_related:
        return '—'
    if score >= 6:
        return '高'
    if score >= 4:
        return '中'
    return '低'


def analyze(title, abstract, d):
    """通用分析函数，d 为领域配置 dict"""
    text   = (title or '') + ' ' + (abstract or '')
    text_l = text.lower()
    title_l = (title or '').lower()

    t1        = find_hits(text_l, d['TIER1'])
    t2_direct = find_hits(text_l, d['TIER2'])
    t2_gated  = find_hits(text_l, d['TIER2_GATED']) if has_context(text_l, d['CONTEXT']) else []
    t2        = t2_direct + t2_gated
    t3        = find_hits(text_l, d['TIER3'])
    fps       = detect_false_positives(text_l, d['FALSE_POSITIVE_PATTERNS'])
    fp_count  = len(fps)
    dn        = d['name']

    if t1:
        is_related = True
        kws  = list(dict.fromkeys([h[0] for h in t1]))[:4]
        snip = best_snippet(t1)
        reason = ('标题/摘要出现核心{}术语【'.format(dn) + '、'.join(kws) + '】，'
                  '直接涉及{}。'.format(dn) + '证据片段："' + snip + '"')
    elif t2:
        is_related = True
        kws  = list(dict.fromkeys([h[0] for h in t2]))[:4]
        snip = best_snippet(t2)
        reason = ('涉及{}特异性分子/机制【'.format(dn) + '、'.join(kws) + '】，'
                  '内容与{}直接相关。'.format(dn) + '证据片段："' + snip + '"')
    elif t3:
        unique3 = list(dict.fromkeys([h[0] for h in t3]))
        if len(unique3) >= 2 and fp_count == 0:
            is_related = True
            reason = ('摘要多次提及{}关联词汇【'.format(dn) + '、'.join(unique3[:4]) + '】，'
                      '且无其他领域干扰，综合判断相关。')
        elif len(unique3) >= 2 and fp_count > 0:
            is_related = False
            fp_names = '、'.join([f[0] for f in fps])
            reason = ('出现{}关联词汇【'.format(dn) + '、'.join(unique3[:3]) + '】，'
                      '但同时检测到干扰词【' + fp_names + '】，综合判断为不相关。')
        else:
            is_related = False
            w = unique3[0] if unique3 else ''
            reason = ('仅出现一般性词汇【' + w + '】，未见{}核心术语，不属于本研究领域。'.format(dn))
    else:
        is_related = False
        topic_kws = re.findall(r'\b[A-Z][A-Za-z0-9\-]{3,}\b', (title or '')[:200])
        topic = '、'.join(list(dict.fromkeys(topic_kws))[:5]) or '（无明显专业词）'
        reason = ('未检测到任何{}相关术语。研究核心关键词：'.format(dn) + topic
                  + '，与{}领域无关。'.format(dn))

    score      = calc_score(t1, t2, t3, title_l, fp_count, is_related)
    conclusion = '相关' if is_related else '不相关'
    level      = score_to_level(score, is_related)
    category   = classify_category(text_l, is_related, d['CATEGORY_PATTERNS'], d['fallback_label'])
    species    = detect_species(text_l, d['SPECIES_PATTERNS'])
    return conclusion, level, score, category, species, reason


def run_domain(d, src_path=SRC_PATH):
    """读取 src_path，按领域 d 分析，输出到 d['out_path']"""
    wb_in = openpyxl.load_workbook(src_path)
    ws_in = wb_in.active

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = d['sheet_title']

    ws_out.append(['title', 'abstract', 'source_id',
                   d['conclusion_col'], '相关性强度', '相关性评分(0-10)',
                   '研究子类别', '涉及物种/研究类型', '判断依据'])

    stats = {
        'total': 0, 'related': 0,
        'level':      defaultdict(int),
        'category':   defaultdict(int),
        'species':    defaultdict(int),
        'score_dist': defaultdict(int),
    }

    for row in ws_in.iter_rows(min_row=2, values_only=True):
        title    = str(row[0]) if row[0] else ''
        abstract = str(row[1]) if row[1] else ''
        source   = row[2] if len(row) > 2 else ''

        conclusion, level, score, category, species, reason = analyze(title, abstract, d)
        ws_out.append([title, abstract, source,
                       conclusion, level, score, category, species, reason])

        stats['total'] += 1
        if conclusion == '相关':
            stats['related'] += 1
            stats['level'][level] += 1
            for cat in category.split(' / '):
                stats['category'][cat] += 1
        for sp in species.split('、'):
            if sp and sp != '未明确':
                stats['species'][sp] += 1
        stats['score_dist'][score] += 1

    # ── 分析结果 Sheet 样式 ──────────────────────────────────────────
    cw = d.get('col_widths', [55, 75, 15, 14, 10, 16, 30, 22, 88])
    for i, w in enumerate(cw, 1):
        ws_out.column_dimensions[get_column_letter(i)].width = w

    hc = d['header_color']
    hdr_fill = PatternFill(start_color=hc, end_color=hc, fill_type='solid')
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    for cell in ws_out[1]:
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws_out.row_dimensions[1].height = 36

    green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    red_fill   = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    high_fill  = PatternFill(start_color='375623', end_color='375623', fill_type='solid')
    mid_fill   = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
    low_fill   = PatternFill(start_color='F4B942', end_color='F4B942', fill_type='solid')
    gray_fill  = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

    for row in ws_out.iter_rows(min_row=2):
        concl = row[3].value
        lv    = row[4].value
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        row[3].fill = green_fill if concl == '相关' else red_fill
        row[3].font = Font(bold=True)
        if   lv == '高': row[4].fill = high_fill; row[4].font = Font(bold=True, color='FFFFFF')
        elif lv == '中': row[4].fill = mid_fill;  row[4].font = Font(bold=True)
        elif lv == '低': row[4].fill = low_fill;  row[4].font = Font(bold=True)
        else:             row[4].fill = gray_fill

    ws_out.freeze_panes = 'A2'

    # ── 统计汇总 Sheet ───────────────────────────────────────────────
    ws_stat = wb_out.create_sheet('统计汇总')
    sc = d.get('stat_color', hc)
    stripe_c = d.get('stripe_color', 'D9E1F2')
    STAT_FILL   = PatternFill(start_color=sc,       end_color=sc,       fill_type='solid')
    STRIPE_FILL = PatternFill(start_color=stripe_c, end_color=stripe_c, fill_type='solid')

    def write_section(ws, sec_title, data_rows, row_start):
        ws.merge_cells(start_row=row_start, start_column=1,
                       end_row=row_start, end_column=2)
        c = ws.cell(row_start, 1, sec_title)
        c.font      = Font(bold=True, color='FFFFFF', size=11)
        c.fill      = STAT_FILL
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row_start].height = 22
        for i, (k, v) in enumerate(data_rows, 1):
            kc = ws.cell(row_start + i, 1, k)
            vc = ws.cell(row_start + i, 2, v)
            kc.alignment = Alignment(horizontal='left',   vertical='center')
            vc.alignment = Alignment(horizontal='center', vertical='center')
            if i % 2 == 0:
                kc.fill = STRIPE_FILL
                vc.fill = STRIPE_FILL
        return row_start + len(data_rows) + 2

    total   = stats['total']
    related = stats['related']
    dn      = d['name']

    r = write_section(ws_stat, '一、总体统计', [
        ('总文献数',           total),
        ('{}相关'.format(dn),  related),
        ('非{}相关'.format(dn), total - related),
        ('相关占比',           '{:.1f}%'.format(related / total * 100 if total else 0)),
    ], 1)
    r = write_section(ws_stat, '二、相关性强度分布（相关文献中）', [
        ('高度相关（6-10分）', stats['level'].get('高', 0)),
        ('中度相关（4-5分）',  stats['level'].get('中', 0)),
        ('低度相关（2-3分）',  stats['level'].get('低', 0)),
    ], r)
    r = write_section(ws_stat, '三、研究子类别分布', [
        (k, v) for k, v in sorted(stats['category'].items(), key=lambda x: -x[1])
    ], r)
    r = write_section(ws_stat, '四、涉及物种 / 研究类型分布（所有文献）', [
        (k, v) for k, v in sorted(stats['species'].items(), key=lambda x: -x[1])
    ], r)
    r = write_section(ws_stat, '五、相关性评分分布（0-10分，所有文献）', [
        ('得分 {}'.format(s), cnt)
        for s, cnt in sorted(stats['score_dist'].items(), key=lambda x: -x[0])
        if cnt > 0
    ], r)

    ws_stat.column_dimensions['A'].width = 34
    ws_stat.column_dimensions['B'].width = 18
    ws_stat.sheet_view.showGridLines = True

    wb_out.save(d['out_path'])

    print('=== {} 分析完成 ==='.format(dn))
    print('总文献: {}  相关: {} 篇 ({:.1f}%)'.format(
        total, related, related / total * 100 if total else 0))
    print('相关性强度 -> 高:{} 中:{} 低:{}'.format(
        stats['level'].get('高', 0), stats['level'].get('中', 0), stats['level'].get('低', 0)))
    for k, v in sorted(stats['category'].items(), key=lambda x: -x[1]):
        print('  {}: {} 篇'.format(k, v))
    print('输出: {}\n'.format(d['out_path']))


# ══════════════════════════════════════════════════════════════════════
# 领域配置：COPD
# ══════════════════════════════════════════════════════════════════════
COPD = {
    'name':           'COPD',
    'sheet_title':    'COPD分析',
    'conclusion_col': 'COPD相关结论',
    'fallback_label': 'COPD相关（其他）',
    'header_color':   'C55A11',
    'stat_color':     'C55A11',
    'stripe_color':   'FAE5DC',
    'out_path':       'E:/Project/BioInfo/temp/sets_COPD.xlsx',
    'col_widths':     [55, 75, 15, 14, 10, 16, 30, 22, 88],

    'TIER1': [
        (r'\bCOPD\b',                                   'COPD'),
        (r'\bchronic obstructive pulmonary disease\b',  'chronic obstructive pulmonary disease(COPD)'),
        (r'\bchronic obstructive lung disease\b',       'chronic obstructive lung disease'),
        (r'\bpulmonary emphysema\b',                    'pulmonary emphysema'),
        (r'\bemphysema\b',                              'emphysema'),
        (r'\bchronic bronchitis\b',                     'chronic bronchitis'),
        (r'\balpha.1.antitrypsin deficiency\b',         'alpha-1-antitrypsin deficiency(AATD)'),
        (r'\bAAT deficiency\b',                         'AAT deficiency(AATD)'),
        (r'\bAATD\b',                                   'AATD'),
        (r'\bAECOPD\b',                                 'AECOPD'),
        (r'\bacute exacerbation.{0,20}(?:COPD|chronic obstructive)\b',
                                                        'acute exacerbation of COPD(AECOPD)'),
        (r'\bCOPD.{0,25}exacerbation\b',               'COPD exacerbation'),
        (r'\bairflow obstruction\b',                    'airflow obstruction'),
        (r'\bairflow limitation\b',                     'airflow limitation'),
        (r'\bchronic airflow\b',                        'chronic airflow'),
        (r'\bsmall airway disease\b',                   'small airway disease'),
        (r'\bsmall airway dysfunction\b',               'small airway dysfunction'),
        (r'\bair trapping\b',                           'air trapping'),
        (r'\blung hyperinflation\b',                    'lung hyperinflation'),
        (r'\bpulmonary hyperinflation\b',               'pulmonary hyperinflation'),
        (r'\bdynamic hyperinflation\b',                 'dynamic hyperinflation'),
    ],

    'TIER2': [
        (r'\bFEV1\s*/\s*FVC\b',                         'FEV1/FVC ratio'),
        (r'\bFEV1\.?/\.?FVC\b',                         'FEV1/FVC ratio'),
        (r'\bGOLD stage[s]?\b',                         'GOLD stage'),
        (r'\bGOLD grade[s]?\b',                         'GOLD grade'),
        (r'\bGOLD classification\b',                    'GOLD classification'),
        (r'\bGOLD criteri\w+\b',                        'GOLD criteria'),
        (r'\bGOLD group[s]?\b',                         'GOLD group'),
        (r'\bCAT score\b',                              'CAT score(COPD Assessment Test)'),
        (r'\bCOPD Assessment Test\b',                   'COPD Assessment Test(CAT)'),
        (r'\bSt\.? ?George.{0,5}Respiratory Questionnaire\b', 'SGRQ'),
        (r'\bSGRQ\b',                                   'SGRQ'),
        (r'\bpack.year[s]?\b',                          'pack-year(s)'),
        (r'\bpack year[s]?\b',                          'pack-year(s)'),
        (r'\bmMRC\b',                                   'mMRC dyspnea scale'),
        (r'\bmodified medical research council\b',      'Modified Medical Research Council(mMRC)'),
        (r'\b6.minute walk\b',                          '6-minute walk test'),
        (r'\b6MWT\b',                                   '6MWT'),
        (r'\btiotropium\b',                             'tiotropium(LAMA)'),
        (r'\bumeclidinium\b',                           'umeclidinium(LAMA)'),
        (r'\bglycopyrronium\b',                         'glycopyrronium(LAMA)'),
        (r'\baclidinium\b',                             'aclidinium(LAMA)'),
        (r'\broflumilast\b',                            'roflumilast(PDE4 inhibitor)'),
        (r'\bindacaterol\b',                            'indacaterol(LABA)'),
        (r'\bolodaterol\b',                             'olodaterol(LABA)'),
        (r'\bvilanterol\b',                             'vilanterol(LABA)'),
        (r'\bLAMA\b',                                   'LAMA(long-acting muscarinic antagonist)'),
        (r'\bphosphodiesterase.4 inhibitor\b',          'PDE4 inhibitor'),
        (r'\bPDE4 inhibitor\b',                         'PDE4 inhibitor'),
        (r'\bpulmonary rehabilitation\b',               'pulmonary rehabilitation'),
        (r'\bsputum neutrophil\w*\b',                   'sputum neutrophil'),
        (r'\bsputum eosinophil\w*\b',                   'sputum eosinophil'),
        (r'\bsputum biomarker[s]?\b',                   'sputum biomarker'),
        (r'\bsputum microbiome\b',                      'sputum microbiome'),
        (r'\bperipheral blood eosinophil[s]?\b',        'peripheral blood eosinophil'),
    ],

    'TIER2_GATED': [
        (r'\bneutrophil elastase\b',                    'neutrophil elastase(NE)'),
        (r'\bMMP.?12\b',                               'MMP-12(macrophage elastase)'),
        (r'\bmatrix metalloproteinase.?12\b',           'matrix metalloproteinase-12'),
        (r'\bMMP.?9\b',                                'MMP-9'),
        (r'\bmatrix metalloproteinase.?9\b',            'matrix metalloproteinase-9'),
        (r'\bSERPINA1\b',                              'SERPINA1(alpha-1 antitrypsin gene)'),
        (r'\balpha.1.antitrypsin\b',                    'alpha-1-antitrypsin(AAT)'),
        (r'\bMUC5[AB]\b',                               'MUC5A/B(airway mucin)'),
        (r'\bmucociliary clearance\b',                  'mucociliary clearance'),
        (r'\bmucociliary dysfunction\b',                'mucociliary dysfunction'),
        (r'\bgoblet cell hyperplasia\b',                'goblet cell hyperplasia'),
        (r'\bgoblet cell metaplasia\b',                 'goblet cell metaplasia'),
        (r'\bFEV1\b',                                   'FEV1'),
        (r'\bFVC\b',                                    'FVC'),
        (r'\bdyspnea\b',                                'dyspnea'),
        (r'\bdyspnoea\b',                               'dyspnoea'),
        (r'\bbreathlessness\b',                         'breathlessness'),
        (r'\bexacerbation[s]?\b',                       'exacerbation'),
        (r'\bICS\b',                                    'ICS(inhaled corticosteroid)'),
        (r'\binhaled corticosteroid[s]?\b',             'inhaled corticosteroid(ICS)'),
        (r'\bbronchodilator[s]?\b',                     'bronchodilator'),
        (r'\bLABA\b',                                   'LABA(long-acting beta-agonist)'),
        (r'\bsalmeterol\b',                             'salmeterol(LABA)'),
        (r'\bformoterol\b',                             'formoterol(LABA)'),
        (r'\bspirometry\b',                             'spirometry'),
        (r'\bazithromycin\b',                           'azithromycin'),
        (r'\boxidative stress\b',                       'oxidative stress'),
        (r'\bneutrophil[s]?\b',                         'neutrophil'),
        (r'\binterleukin.8\b',                          'IL-8(CXCL8)'),
        (r'\bIL.8\b',                                   'IL-8'),
        (r'\bCXCL8\b',                                  'CXCL8(IL-8)'),
        (r'\bTNF.alpha\b',                              'TNF-alpha'),
        (r'\bTNF.\u03b1\b',                             'TNF-\u03b1'),
        (r'\bC.reactive protein\b',                     'C-reactive protein(CRP)'),
        (r'\bCRP\b',                                    'CRP(C-reactive protein)'),
        (r'\blung function\b',                          'lung function'),
        (r'\brespiratory failure\b',                    'respiratory failure'),
        (r'\bhyperinflation\b',                         'hyperinflation'),
        (r'\bmucus hypersecretion\b',                   'mucus hypersecretion'),
    ],

    'CONTEXT': [
        r'\bCOPD\b', r'\bchronic obstructive\b', r'\bemphysema\b',
        r'\bchronic bronchitis\b', r'\bairflow obstruction\b',
        r'\bairflow limitation\b', r'\bAECOPD\b',
        r'\bGOLD stage\b', r'\bGOLD grade\b', r'\bGOLD classification\b',
        r'\btiotropium\b', r'\broflumilast\b',
        r'\bpack.year\b', r'\bpack year\b',
        r'\bFEV1\s*/\s*FVC\b', r'\bFEV1\.?/\.?FVC\b',
        r'\bAATD\b', r'\bsmall airway\b', r'\bLAMA\b',
        r'\bpulmonary rehabilitation\b', r'\blung hyperinflation\b',
        r'\bair trapping\b',
    ],

    'TIER3': [
        (r'\brespiratory\b',         'respiratory'),
        (r'\bpulmonary\b',           'pulmonary'),
        (r'\bbronchial\b',           'bronchial'),
        (r'\blung[s]?\b',            'lung'),
        (r'\bcough\b',               'cough'),
        (r'\bsputum\b',              'sputum'),
        (r'\bmucus\b',               'mucus'),
        (r'\bairway[s]?\b',          'airway'),
        (r'\bwheez\w+\b',            'wheeze/wheezing'),
        (r'\bbronchi\w*\b',          'bronchi'),
    ],

    'FALSE_POSITIVE_PATTERNS': [
        (r'\bidiopathic pulmonary fibrosis\b',    'idiopathic pulmonary fibrosis(IPF)'),
        (r'\bcystic fibrosis\b',                  'cystic fibrosis（囊性纤维化）'),
        (r'\bacute respiratory distress syndrome\b','ARDS（急性呼吸窘迫综合征）'),
        (r'\blung cancer\b',                      'lung cancer（肺癌）'),
        (r'\blung carcinoma\b',                   'lung carcinoma（肺癌）'),
        (r'\bnon.small cell lung\b',              'NSCLC（非小细胞肺癌）'),
        (r'\bsmall cell lung\b',                  'SCLC（小细胞肺癌）'),
        (r'\bpulmonary adenocarcinoma\b',         'pulmonary adenocarcinoma（肺腺癌）'),
        (r'\bpulmonary arterial hypertension\b',  'pulmonary arterial hypertension(PAH)'),
        (r'\bbronchiectasis\b',                   'bronchiectasis（支气管扩张症）'),
        (r'\bobstructive sleep apnea\b',          'obstructive sleep apnea(OSA)'),
        (r'\bsleep apnea\b',                      'sleep apnea（睡眠呼吸暂停）'),
        (r'\bCOVID.19\b',                         'COVID-19'),
        (r'\bSARS.CoV.2\b',                       'SARS-CoV-2'),
        (r'\binterstitial lung disease\b',        'interstitial lung disease(ILD)'),
        (r'\bpulmonary fibrosis\b',               'pulmonary fibrosis（肺纤维化）'),
        (r'\basthma\b',                           'asthma（哮喘，非COPD语境）'),
        (r'\bsarcoidosis\b',                      'sarcoidosis（结节病）'),
        (r'\bmesothelioma\b',                     'mesothelioma（间皮瘤）'),
        (r'\bpneumonia\b',                        'pneumonia（肺炎）'),
        (r'\btuberculosis\b',                     'tuberculosis（结核）'),
    ],

    'CATEGORY_PATTERNS': {
        '发病机制研究': [
            r'\bpathogenesis\b', r'\bpathophysiology\b',
            r'\boxidative stress\b', r'\bairway inflammation\b',
            r'\blung inflammation\b', r'\bprotease.antiprotease\b',
            r'\bneutrophil elastase\b', r'\bmatrix metalloproteinase\b',
            r'\bmucus hypersecretion\b', r'\bgoblet cell\b',
            r'\bmucociliary\b', r'\bairway remodel\w+\b',
            r'\bparenchymal destruction\b', r'\blungs? destroy\w+\b',
            r'\bprotease.{0,20}inhibitor\b',
            r'\bsenescen\w+\b', r'\bcellular senescen\w+\b',
        ],
        '诊断与肺功能': [
            r'\bspirometry\b', r'\blung function\b', r'\bpulmonary function\b',
            r'\bFEV1\b', r'\bFVC\b', r'\bFEV1.?/.?FVC\b',
            r'\bGOLD stage\b', r'\bGOLD grade\b',
            r'\bcomputed tomography\b', r'\b(?<!\w)CT scan\b',
            r'\bquantitative CT\b',
            r'\bbiomarker.{0,30}(?:diagnos|detect)\w*\b',
            r'\bdiagnos\w+.{0,30}COPD\b', r'\blung densitometry\b',
        ],
        '治疗与药物研究': [
            r'\bclinical trial\b', r'\brandomized.{0,15}trial\b',
            r'\btiotropium\b', r'\broflumilast\b',
            r'\bLAMA\b', r'\bLABA\b',
            r'\binhaled corticosteroid\b', r'\bICS\b',
            r'\bbronchodilator\b', r'\bphosphodiesterase.4\b', r'\bPDE4\b',
            r'\bdisease.modifying\b', r'\bindacaterol\b', r'\bumeclidinium\b',
            r'\bglycopyrronium\b', r'\baclidinium\b',
            r'\bpharmacolog\w+\b', r'\bgene therapy\b',
        ],
        '急性加重(AECOPD)': [
            r'\bAECOPD\b', r'\bacute exacerbation\b',
            r'\bCOPD.{0,25}exacerbation\b', r'\bexacerbation.{0,25}COPD\b',
            r'\bhospitalization\b', r'\bhospital admission\b',
            r'\binfection.{0,30}exacerbation\b', r'\bviral exacerbation\b',
            r'\bbacterial.{0,20}(?:exacerbation|infection)\b',
            r'\bexacerbation.{0,30}(?:risk|frequency|rate|prevention)\b',
        ],
        '流行病学/临床研究': [
            r'\bprevalence\b', r'\bincidence\b', r'\bcohort\b',
            r'\bcase.control\b', r'\bprospective\b',
            r'\bepidemiolog\w+\b', r'\bpopulation.based\b',
            r'\brisk factor[s]?\b', r'\bmortality\b',
            r'\bprognosis\b', r'\bsurvival\b', r'\bnatural history\b',
        ],
        '肺康复/运动耐力': [
            r'\bpulmonary rehabilitation\b',
            r'\bexercise capacity\b', r'\bexercise tolerance\b',
            r'\b6.minute walk\b', r'\b6MWT\b',
            r'\bphysical activity\b', r'\bphysical performance\b',
            r'\bmuscle strength\b', r'\bskeletal muscle\b',
            r'\bcachexia\b', r'\bsarcopenia\b',
            r'\bendurance train\w+\b', r'\bresistance train\w+\b',
            r'\binspiratory muscle\b',
        ],
        '生物标志物研究': [
            r'\bbiomarker[s]?\b', r'\bsputum biomarker\b',
            r'\bblood biomarker\b', r'\bplasma biomarker\b',
            r'\bserum.{0,20}(?:marker|level|concentrat)\w*\b',
            r'\bperipheral blood\b', r'\bvolatile organic compound[s]?\b',
            r'\bproteomic\w*\b', r'\bmetabolomic\w*\b',
            r'\bgene expression\b', r'\bblood eosinophil\b',
        ],
        '动物/细胞模型': [
            r'\bmouse model\b', r'\bmurine model\b',
            r'\bcigarette smoke.{0,30}(?:model|expos|induc)\w*\b',
            r'\bsmoke.{0,20}expos\w+\b', r'\brat model\b', r'\bguinea pig\b',
            r'\bin vitro\b', r'\bcell.?based\b', r'\bcell culture\b',
            r'\bprimary cell[s]?\b', r'\bairway epithelial\b',
            r'\bbronchial epithelial\b',
        ],
        '合并症研究': [
            r'\bcomorbidity\b', r'\bcomorbidities\b',
            r'\bcardiovascular\b', r'\bheart failure\b',
            r'\bosteoporosis\b', r'\banxiety\b', r'\bdepression\b',
            r'\bdiabetes\b', r'\bmetabolic syndrome\b',
            r'\bCOPD.{0,30}cancer\b', r'\bcancer.{0,30}COPD\b',
            r'\banemia\b', r'\bmalnutrition\b', r'\bsystemic inflammation\b',
        ],
    },

    'SPECIES_PATTERNS': {
        '人类/临床': [r'\bhuman[s]?\b', r'\bpatient[s]?\b', r'\bclinical\b',
                      r'\bHomo sapiens\b', r'\bhuman lung\b'],
        '小鼠/大鼠': [r'\bmouse\b', r'\bmice\b', r'\bmurine\b', r'\brat\b',
                      r'\bMus musculus\b'],
        '大动物/其他': [r'\bguinea pig\b', r'\bsheep\b', r'\bporcine\b',
                        r'\bpig model\b'],
        '细胞系/体外': [r'\bcell line[s]?\b', r'\bin vitro\b',
                        r'\bprimary cell[s]?\b', r'\bairway epithelial\b',
                        r'\borganoid\b'],
        '计算/综述': [r'\bin silico\b', r'\bcomputational\b',
                      r'\breview\b', r'\bmeta.analysis\b', r'\bsystematic review\b'],
    },
}


# ══════════════════════════════════════════════════════════════════════
# 领域配置：神经退行性疾病
# ══════════════════════════════════════════════════════════════════════
NEURO = {
    'name':           '神经退行性疾病',
    'sheet_title':    '神经退行性疾病分析',
    'conclusion_col': '神经退行性疾病相关结论',
    'fallback_label': '神经退行性疾病相关（其他）',
    'header_color':   '4B1C82',
    'stat_color':     '4B1C82',
    'stripe_color':   'EAE0F5',
    'out_path':       'E:/Project/BioInfo/temp/sets_NeuroDegenerative.xlsx',
    'col_widths':     [55, 75, 15, 16, 10, 16, 30, 22, 88],

    'TIER1': [
        (r'\bneurodegenerative disease[s]?\b',        'neurodegenerative disease'),
        (r'\bneurodegenerative disorder[s]?\b',        'neurodegenerative disorder'),
        (r'\bneurodegeneration\b',                     'neurodegeneration'),
        (r"\balzheimer'?s?\s*disease\b",               "Alzheimer's disease(AD)"),
        (r'\balzheimer\s*disease\b',                   'Alzheimer disease(AD)'),
        (r"\bparkinson'?s?\s*disease\b",               "Parkinson's disease(PD)"),
        (r'\bparkinson\s*disease\b',                   'Parkinson disease(PD)'),
        (r'\bamyotrophic lateral sclerosis\b',         'amyotrophic lateral sclerosis(ALS)'),
        (r"\bhuntington'?s?\s*disease\b",              "Huntington's disease(HD)"),
        (r'\bhuntington\s*disease\b',                  'Huntington disease(HD)'),
        (r'\bfrontotemporal dementia\b',               'frontotemporal dementia(FTD)'),
        (r'\bfrontotemporal lobar degeneration\b',     'FTLD'),
        (r'\bmultiple sclerosis\b',                    'multiple sclerosis(MS)'),
        (r'\blewy body dementia\b',                    'Lewy body dementia(LBD)'),
        (r'\bdementia with lewy bodies\b',             'dementia with Lewy bodies(DLB)'),
        (r'\bspinocerebellar ataxia\b',                'spinocerebellar ataxia(SCA)'),
        (r'\bprogressive supranuclear palsy\b',        'progressive supranuclear palsy(PSP)'),
        (r'\bcorticobasal degeneration\b',             'corticobasal degeneration(CBD)'),
        (r'\bmotor neuron disease[s]?\b',              'motor neuron disease(MND)'),
        (r'\bprion disease[s]?\b',                     'prion disease'),
        (r'\bCreutzfeldt.Jakob disease\b',             'Creutzfeldt-Jakob disease(CJD)'),
        (r'\bvascular dementia\b',                     'vascular dementia'),
        (r'\bmixed dementia\b',                        'mixed dementia'),
        (r'\bprimary progressive aphasia\b',           'primary progressive aphasia(PPA)'),
        (r'\bmultiple system atrophy\b',               'multiple system atrophy(MSA)'),
        (r'\bneuromyelitis optica\b',                  'neuromyelitis optica(NMO)'),
    ],

    'TIER2': [
        (r'\bamyloid[- ]?beta\b',               'amyloid-beta(Aβ)'),
        (r'\bamyloid beta\b',                   'amyloid beta(Aβ)'),
        (r'\bA\s*[Ββ]\d{0,2}\b',               'Aβ'),
        (r'\bneurofibrillary tangle[s]?\b',     'neurofibrillary tangles(NFT)'),
        (r'\bamyloid plaque[s]?\b',             'amyloid plaques'),
        (r'\bsenile plaque[s]?\b',              'senile plaques'),
        (r'\b(?<!\w)APP\b',                     'APP(amyloid precursor protein)'),
        (r'\bamyloid precursor protein\b',      'amyloid precursor protein(APP)'),
        (r'\bBACE[12]?\b',                      'BACE(β-secretase)'),
        (r'\bbeta.secretase\b',                 'beta-secretase(BACE)'),
        (r'\bgamma.secretase\b',                'gamma-secretase'),
        (r'\bPSEN[12]\b',                       'PSEN(presenilin)'),
        (r'\bpresenilin[s\-]?\d?\b',            'presenilin'),
        (r'\bapolipoprotein E\b',               'ApoE'),
        (r'\b(?<!\w)APOE\b',                    'APOE'),
        (r'\bAPOE.?[e\u03b5]?4\b',             'APOE4'),
        (r'\btau protein[s]?\b',                'tau protein'),
        (r'\bphospho.?tau\b',                   'phospho-tau(p-tau)'),
        (r'\bhyperphosphorylated tau\b',        'hyperphosphorylated tau'),
        (r'\b(?<!\w)MAPT\b',                    'MAPT(tau gene)'),
        (r'\btauopathy\b',                      'tauopathy'),
        (r'\balpha.synuclein\b',                'alpha-synuclein(α-syn)'),
        (r'\b\u03b1.synuclein\b',               'α-synuclein'),
        (r'\bsynucleinopathy\b',               'synucleinopathy'),
        (r'\b(?<!\w)SNCA\b',                    'SNCA'),
        (r'\bLewy bod(?:y|ies)\b',              'Lewy body/bodies'),
        (r'\b(?<!\w)LRRK2\b',                   'LRRK2'),
        (r'\b(?<!\w)PINK1\b',                   'PINK1'),
        (r'\bParkin\b',                         'Parkin/PRKN'),
        (r'\b(?<!\w)PRKN\b',                    'PRKN'),
        (r'\bDJ.1\b',                           'DJ-1(PARK7)'),
        (r'\bsubstantia nigra\b',               'substantia nigra'),
        (r'\bdopaminergic neuron[s]?\b',        'dopaminergic neurons'),
        (r'\bdopamine.{0,15}(?:neuron|loss|deficit|depletion)\w*\b', 'dopamine neuron loss'),
        (r'\bTDP.43\b',                         'TDP-43(TARDBP)'),
        (r'\b(?<!\w)TARDBP\b',                  'TARDBP'),
        (r'\bFUS protein\b',                    'FUS protein'),
        (r'\b(?<!\w)C9orf72\b',                 'C9orf72'),
        (r'\bhexanucleotide repeat expansion\b','hexanucleotide repeat(C9orf72)'),
        (r'\b(?<!\w)SOD1\b',                    'SOD1'),
        (r'\bhuntingtin\b',                     'huntingtin(HTT)'),
        (r'\b(?<!\w)HTT\b',                     'HTT'),
        (r'\bCAG repeat[s]?\b',                 'CAG repeat'),
        (r'\bpolyglutamine\b',                  'polyglutamine(polyQ)'),
        (r'\bpoly.?Q\b',                        'polyQ'),
        (r'\bstriatum\b',                       'striatum'),
        (r'\bstriatal\b',                       'striatal'),
        (r'\bprion protein\b',                  'prion protein(PrP)'),
        (r'\b(?<!\w)PRNP\b',                    'PRNP'),
        (r'\bPrP[CSc]{0,2}\b',                  'PrP'),
        (r'\bprion.like propagation\b',         'prion-like propagation'),
        (r'\bprion.like spread\b',              'prion-like spread'),
        (r'\bneuroinflammation\b',              'neuroinflammation'),
        (r'\bmicroglial? activat\w+\b',         'microglial activation'),
        (r'\bneuronal death\b',                 'neuronal death'),
        (r'\bneurotoxicity\b',                  'neurotoxicity'),
        (r'\bneuronal loss\b',                  'neuronal loss'),
        (r'\bcognitive impairment\b',           'cognitive impairment'),
        (r'\bcognitive decline\b',              'cognitive decline'),
        (r'\bmild cognitive impairment\b',      'mild cognitive impairment(MCI)'),
        (r'\b(?<!\w)MCI\b',                     'MCI'),
        (r'\bmemory impairment\b',              'memory impairment'),
        (r'\bneurofilament light\b',            'neurofilament light chain(NfL)'),
        (r'\b(?<!\w)NfL\b',                     'NfL'),
        (r'\bCSF biomarker[s]?\b',              'CSF biomarker'),
        (r'\bblood biomarker[s]?\b',            'blood biomarker'),
        (r'\bplasma biomarker[s]?\b',           'plasma biomarker'),
        (r'\bGFAP.{0,20}(?:biomarker|plasma|CSF|serum)\b', 'GFAP biomarker'),
        (r'\blecanemab\b',                      'lecanemab'),
        (r'\bdonanemab\b',                      'donanemab'),
        (r'\baducanumab\b',                     'aducanumab'),
        (r'\bnusinersen\b',                     'nusinersen'),
        (r'\briluzole\b',                       'riluzole(ALS)'),
        (r'\bedaravone\b',                      'edaravone(ALS)'),
        (r'\bdeep brain stimulation\b',         'deep brain stimulation(DBS)'),
    ],

    'TIER2_GATED': [
        (r'\b(?<!\w)MS(?!\w)\b', 'MS(multiple sclerosis)'),
        (r'\bdementia\b',         'dementia'),
        (r'\bprotein aggregat\w+\b',            'protein aggregation'),
        (r'\bprotein misfolding\b',             'protein misfolding'),
        (r'\bproteinopathy\b',                  'proteinopathy'),
        (r'\bmitochondrial dysfunction\b',      'mitochondrial dysfunction'),
        (r'\bautophagy.{0,20}(?:impair|dysfunc|defect)\w*\b', 'autophagy impairment'),
        (r'\bubiquitin.proteasome system\b',    'ubiquitin-proteasome system'),
        (r'\baxonal degeneration\b',            'axonal degeneration'),
        (r'\bexcitotoxicity\b',                 'excitotoxicity'),
        (r'\bsynaptic loss\b',                  'synaptic loss'),
        (r'\bsynaptic dysfunction\b',           'synaptic dysfunction'),
        (r'\b(?<!\w)GBA\b',                     'GBA'),
        (r'\bglucocerebrosidase\b',             'glucocerebrosidase(GBA)'),
        (r'\bmemory loss\b',                    'memory loss'),
    ],

    'CONTEXT': [
        r'\bneurodegener\w+\b', r'\bneurolog\w+\b', r'\bneuron\w*\b',
        r'\bcognitive\b', r'\bneuropathol\w+\b', r'\b(?<!\w)CNS\b',
        r'\bcerebral\b', r'\bhippocampus\b', r'\bsubstantia nigra\b',
        r'\bdopamin\w+\b', r'\bamyloid\b', r'\btau\b', r'\bsynuclein\b',
        r'\bmotor neuron\b', r'\bparkinson\w*\b', r'\balzheimer\w*\b',
        r'\bbasal ganglia\b', r'\bneurotoxic\w+\b',
        r'\bcognitive impairment\b', r'\bmemory impairment\b',
        r'\bbrain atrophy\b', r'\bLewy bod\w+\b', r'\bneurofibrill\w+\b',
    ],

    'TIER3': [
        (r'\bneuron[s]?\b',              'neuron'),
        (r'\bneuronal\b',                'neuronal'),
        (r'\baxon[s]?\b',                'axon'),
        (r'\bsynapse[s]?\b',             'synapse'),
        (r'\bsynaptic\b',                'synaptic'),
        (r'\bdopamine\b',                'dopamine'),
        (r'\bmicroglia[l]?\b',           'microglia'),
        (r'\bastrocyte[s]?\b',           'astrocyte'),
        (r'\bcognitive\b',               'cognitive'),
        (r'\bneuroprotect\w+\b',         'neuroprotection'),
        (r'\bneuropatholog\w+\b',        'neuropathology'),
        (r'\bhippocampal\b',             'hippocampal'),
        (r'\bhippocampus\b',             'hippocampus'),
        (r'\bcortical\b',                'cortical'),
        (r'\bcerebral\b',                'cerebral'),
    ],

    'FALSE_POSITIVE_PATTERNS': [
        (r'\bglioblastoma\b',              'glioblastoma（脑肿瘤）'),
        (r'\bglioma\b',                    'glioma（脑肿瘤）'),
        (r'\bmeningioma\b',                'meningioma（脑膜瘤）'),
        (r'\bbrain tumor[s]?\b',           'brain tumor（脑肿瘤）'),
        (r'\bautism spectrum\b',           'autism spectrum disorder（神经发育障碍）'),
        (r'\bschizophrenia\b',             'schizophrenia（精神分裂症）'),
        (r'\bbipolar disorder\b',          'bipolar disorder（双相障碍）'),
        (r'\bdepressive disorder\b',       'depressive disorder（抑郁症）'),
        (r'\bepilep[st]\w+\b',             'epilepsy（癫痫）'),
        (r'\bcerebral ischemi\w+\b',       'cerebral ischemia（脑缺血）'),
        (r'\bischemic stroke\b',           'ischemic stroke（缺血性卒中）'),
        (r'\bspinal cord injur\w+\b',      'spinal cord injury（脊髓损伤）'),
        (r'\bneurodevelopmental\b',        'neurodevelopmental disorder（神经发育障碍）'),
    ],

    'CATEGORY_PATTERNS': {
        '发病机制研究': [
            r'\bpathogenesis\b', r'\bpathophysiology\b',
            r'\bprotein aggregat\w+\b', r'\bprotein misfolding\b',
            r'\bproteinopathy\b', r'\bneuroinflammation\b',
            r'\bmitochondrial dysfunction\b', r'\bsynaptic dysfunction\b',
            r'\bautophagy\b', r'\bneurotoxicity\b',
            r'\bprion.like\b', r'\btauopathy\b', r'\bsynucleinopathy\b',
            r'\bexcitotoxicity\b', r'\baxonal degeneration\b',
        ],
        '生物标志物研究': [
            r'\bbiomarker[s]?\b', r'\bdiagnostic marker[s]?\b',
            r'\b(?<!\w)CSF\b', r'\bplasma biomarker\b', r'\bblood biomarker\b',
            r'\bneurofilament light\b', r'\b(?<!\w)NfL\b',
            r'\bphospho.?tau\b', r'\bPET imaging\b', r'\bPET scan\b',
            r'\bGFAP.{0,20}biomarker\b',
        ],
        '治疗与药物研究': [
            r'\bclinical trial\b',
            r'\bdrug.{0,20}(?:therapy|treatment|target)\b',
            r'\bgene therapy\b', r'\bimmunotherapy\b',
            r'\bstem cell\b', r'\bdeep brain stimulation\b',
            r'\bneuroprotective\b',
            r'\bantisense oligonucleotide\b', r'\bmonoclonal antibody\b',
            r'\blecanemab\b', r'\bdonanemab\b', r'\baducanumab\b',
            r'\bnusinersen\b', r'\briluzole\b', r'\bedaravone\b',
            r'\bphase [I\d]+ (?:trial|study)\b', r'\bdisease.modifying\b',
        ],
        '遗传与基因组学研究': [
            r'\bgenetic.{0,20}(?:risk|variant|mutation|factor)\b',
            r'\b(?<!\w)GWAS\b', r'\bwhole.?exome\b', r'\bwhole.?genome\b',
            r'\bde novo mutation\b', r'\brisk gene[s]?\b',
            r'\bpolymorphism\b', r'\bsingle nucleotide\b',
            r'\bfamilial.{0,20}(?:disease|case|form)\b',
            r'\bmendelian randomization\b', r'\bhaplotype\b',
        ],
        '神经影像学研究': [
            r'\b(?<!\w)MRI\b', r'\bfMRI\b', r'\b(?<!\w)PET\b',
            r'\bneuroimaging\b', r'\bbrain imaging\b',
            r'\bwhite matter\b', r'\bgray matter\b',
            r'\bbrain atrophy\b', r'\bcortical thickness\b',
            r'\bdiffusion tensor\b', r'\b(?<!\w)DTI\b',
        ],
        '流行病学/临床研究': [
            r'\bprevalence\b', r'\bincidence\b', r'\bcohort study\b',
            r'\bcase.control\b', r'\bprospective study\b',
            r'\bepidemiolog\w+\b', r'\bpopulation.based\b',
            r'\bsurvival analysis\b', r'\bclinical feature[s]?\b',
            r'\bclinical manifestation\b', r'\bnatural history\b',
        ],
        '动物模型研究': [
            r'\bmouse model\b', r'\btransgenic mouse\b', r'\bknockout mouse\b',
            r'\brat model\b', r'\bmodel.{0,20}disease\b',
            r'\bDrosophila model\b', r'\bzebrafish model\b',
            r'\bC\. elegans model\b', r'\bin vivo model\b',
            r'\bAAV.{0,20}(?:model|inject|express)\b',
        ],
        '细胞/分子机制研究': [
            r'\bin vitro\b', r'\bcell.based\b', r'\bcell culture\b',
            r'\bneuronal cell\b', r'\binduced pluripotent\b',
            r'\biPSC\b', r'\borganoid\b', r'\bsingle.cell\b',
            r'\bmolecular mechanism\b', r'\bsignaling pathway\b',
        ],
    },

    'SPECIES_PATTERNS': {
        '人类/临床': [r'\bhuman[s]?\b', r'\bpatient[s]?\b', r'\bclinical\b',
                      r'\bHomo sapiens\b', r'\bhuman brain\b'],
        '小鼠/大鼠': [r'\bmouse\b', r'\bmice\b', r'\bmurine\b', r'\brat\b',
                      r'\bMus musculus\b', r'\btransgenic mouse\b'],
        '细胞系/iPSC': [r'\bcell line[s]?\b', r'\bin vitro\b',
                        r'\biPSC\b', r'\binduced pluripotent\b', r'\borganoid\b'],
        '非人灵长类': [r'\bprimate[s]?\b', r'\bmonkey\b', r'\bmacaque\b'],
        '斑马鱼':   [r'\bzebrafish\b', r'\bDanio rerio\b'],
        '果蝇':     [r'\bDrosophila\b', r'\bfruit fly\b'],
        '线虫':     [r'\bC\. elegans\b', r'\bnematode\b'],
        '计算/综述': [r'\bin silico\b', r'\bcomputational\b',
                      r'\breview\b', r'\bmeta.analysis\b', r'\bsystematic review\b'],
    },
}


# ══════════════════════════════════════════════════════════════════════
# 领域配置：RNA 剪接
# ══════════════════════════════════════════════════════════════════════
RNA = {
    'name':           'RNA剪接',
    'sheet_title':    'RNA剪接分析',
    'conclusion_col': 'RNA剪接相关结论',
    'fallback_label': '剪接相关（其他）',
    'header_color':   '1F4E79',
    'stat_color':     '2E75B6',
    'stripe_color':   'D9E1F2',
    'out_path':       'E:/Project/BioInfo/temp/sets_RNAsplicing.xlsx',
    'col_widths':     [55, 75, 15, 14, 10, 16, 30, 22, 88],

    'TIER1': [
        (r'\bRNA splicing\b',               'RNA splicing'),
        (r'\bmRNA splicing\b',              'mRNA splicing'),
        (r'\bpre-?mRNA\b',                  'pre-mRNA'),
        (r'\bsplicin[g]?\b',                'splicing'),
        (r'\bspliceosom\w*\b',              'spliceosome'),
        (r'\balternative splicing\b',       'alternative splicing'),
        (r'\bexon skip+ing\b',              'exon skipping'),
        (r'\bintron retention\b',           'intron retention'),
        (r'\bexon inclusion\b',             'exon inclusion'),
        (r'\bcryptic splicing\b',           'cryptic splicing'),
        (r'\baberrant splicing\b',          'aberrant splicing'),
        (r'\bpathogenic splicing\b',        'pathogenic splicing'),
        (r'\bsplice site[s]?\b',            'splice site'),
        (r'\bsplice junction[s]?\b',        'splice junction'),
        (r'\bsplice variant[s]?\b',         'splice variant'),
        (r'\bsplice isoform[s]?\b',         'splice isoform'),
        (r'\bsplice donor[s]?\b',           'splice donor'),
        (r'\bsplice acceptor[s]?\b',        'splice acceptor'),
        (r"\b5['u2019]?\s*splice\b",        "5' splice site"),
        (r"\b3['u2019]?\s*splice\b",        "3' splice site"),
        (r'\bbranch point[s]?\b',           'branch point'),
        (r'\bintron removal\b',             'intron removal'),
        (r'\bpseudoexon[s]?\b',             'pseudoexon'),
        (r'\bpoison exon[s]?\b',            'poison exon'),
        (r'\bnonsense.mediated decay\b',    'nonsense-mediated decay'),
        (r'\bminor spliceosome\b',          'minor spliceosome'),
        (r'\bU12.type intron[s]?\b',        'U12-type intron'),
        (r'\bGT.AG rule\b',                 'GT-AG rule'),
        (r'\bGU.AG\b',                      'GU-AG'),
        (r'\btrans.splicing\b',             'trans-splicing'),
        (r'\bsplice.switching\b',           'splice-switching'),
        (r'\bsplicing code\b',              'splicing code'),
        (r'\bco.transcriptional splicing\b','co-transcriptional splicing'),
    ],

    'TIER2': [
        (r'\bSF3B[1-4]?\b',                    'SF3B'),
        (r'\bU[12456] snRNA\b',                'snRNA(U1/2/4/5/6)'),
        (r'\bU11 snRNA\b',                     'U11 snRNA'),
        (r'\bU12 snRNA\b',                     'U12 snRNA'),
        (r'\bsnRNP[s]?\b',                     'snRNP'),
        (r'\bU2AF[12]?\b',                     'U2AF'),
        (r'\bSRSF\d*\b',                       'SRSF'),
        (r'\bSR protein[s]?\b',                'SR protein'),
        (r'\bSR.rich\b',                       'SR-rich'),
        (r'\bRBFOX[123]?\b',                   'RBFOX'),
        (r'\bNOVA[12]?\b',                     'NOVA'),
        (r'\bPTBP[123]?\b',                    'PTBP'),
        (r'\bPRPF\d+\b',                       'PRPF'),
        (r'\bMBNL[123]?\b',                    'MBNL'),
        (r'\bTDP-43\b',                        'TDP-43'),
        (r'\bTRA2[AB]?\b',                     'TRA2'),
        (r'\bsplicing factor[s]?\b',           'splicing factor'),
        (r'\bsplicing regul\w+\b',             'splicing regulation'),
        (r'\bsplicing modulator[s]?\b',        'splicing modulator'),
        (r'\bsplicing mutation[s]?\b',         'splicing mutation'),
        (r'\bsplicing defect[s]?\b',           'splicing defect'),
        (r'\bsplicing pattern[s]?\b',          'splicing pattern'),
        (r'\bsplicing efficiency\b',           'splicing efficiency'),
        (r'\bexonic splicing enhancer[s]?\b',  'ESE(exonic splicing enhancer)'),
        (r'\bexonic splicing silencer[s]?\b',  'ESS(exonic splicing silencer)'),
        (r'\bintronic splicing enhancer[s]?\b','ISE(intronic splicing enhancer)'),
        (r'\bintronic splicing silencer[s]?\b','ISS(intronic splicing silencer)'),
        (r'\b(?<!\w)ESE(?!\w)\b',              'ESE'),
        (r'\b(?<!\w)ESS(?!\w)\b',              'ESS'),
        (r'\bpolypyrimidine tract\b',          'polypyrimidine tract'),
        (r'\bsplicing.correcting\b',           'splicing-correcting'),
        (r'\bsplice.correcting\b',             'splice-correcting'),
        (r'\bspliceostatin\b',                 'spliceostatin'),
    ],

    'TIER2_GATED': [
        (r'\b(?<!\w)FUS(?!\w)\b',              'FUS'),
        (r'\b(?<!\w)SON(?!\w)\b',              'SON'),
        (r'\b(?<!\w)NMD(?!\w)\b',              'NMD(nonsense-mediated decay)'),
        (r'\bSMN[12]?\b',                      'SMN1/2'),
        (r'\bspinal muscular atrophy\b',       'spinal muscular atrophy(SMA)'),
        (r'\bantisense oligonucleotide[s]?\b', 'antisense oligonucleotide'),
        (r'\bnusinersen\b',                    'nusinersen'),
        (r'\brisdiplam\b',                     'risdiplam'),
        (r'\bhnRNP\b',                         'hnRNP'),
        (r'\bRBM[0-9]+\b',                     'RBM'),
        (r'\bEIF4A3\b',                        'EIF4A3'),
        (r'\bCLK[1-4]?\b',                     'CLK'),
        (r'\bDYRK[12]?\b',                     'DYRK'),
        (r'\bEWSR1\b',                         'EWSR1'),
    ],

    'CONTEXT': [
        r'\bsplicin[g]?\b', r'\bsplice\b', r'\bspliceosom\w*\b',
        r'\bpre.?mRNA\b', r'\bexon skip\w*\b', r'\bintron retention\b',
        r'\bexon inclusion\b', r'\bSMN2\b', r'\bsnRNP\b', r'\bsnRNA\b',
        r'\bsplice site\b', r'\bbranch point\b', r'\bexon 7\b',
        r'\bpseudoexon\b', r'\bpoison exon\b', r'\bnonsense.mediated decay\b',
        r'\bexon junction complex\b', r'\b(?<!\w)EJC\b',
        r'\b(?<!\w)UPF[123]\b', r'\bpremature termination codon\b',
        r'\bpremature stop codon\b', r'\b(?<!\w)PTC\b', r'\bmRNA surveillance\b',
    ],

    'TIER3': [
        (r'\bexon[s]?\b',               'exon'),
        (r'\bintron[s]?\b',             'intron'),
        (r'\bisoform[s]?\b',            'isoform'),
        (r'\btranscript variant[s]?\b', 'transcript variant'),
        (r'\bmRNA isoform[s]?\b',       'mRNA isoform'),
        (r'\bmRNA processing\b',        'mRNA processing'),
        (r'\bRNA processing\b',         'RNA processing'),
        (r'\bRNA maturation\b',         'RNA maturation'),
        (r'\bpost.transcriptional\b',   'post-transcriptional'),
    ],

    'FALSE_POSITIVE_PATTERNS': [
        (r'\bprotein isoform[s]?\b',        'protein isoform（蛋白层面异构体）'),
        (r'\btranscription factor[s]?\b',   'transcription factor（转录因子）'),
        (r'\bgene transcription\b',         'gene transcription（基因转录）'),
        (r'\btranscriptional regulation\b', 'transcriptional regulation'),
        (r'\btranscriptional activat\w+\b', 'transcriptional activation'),
        (r'\btranscriptional repressor\b',  'transcriptional repressor'),
        (r'\bproteoform[s]?\b',             'proteoform'),
    ],

    'CATEGORY_PATTERNS': {
        '剪接机制研究': [
            r'\bspliceosom\w+\b', r'\bsnRNP\b', r'\bsnRNA\b',
            r'\bbranch point\b',  r'\bsplice site\b',
            r'\bsplicing mechanism\b', r'\bspliceosome assembly\b',
            r'\bGT.AG\b', r'\bGU.AG\b',
        ],
        '疾病相关剪接异常': [
            r'\bpathogenic.{0,30}splicing\b', r'\bsplicing.{0,30}pathogenic\b',
            r'\bdisease.{0,40}splicing\b',    r'\bsplicing.{0,40}disease\b',
            r'\bmutation.{0,30}splicing\b',   r'\bsplicing.{0,30}mutation\b',
            r'\baberrant splicing\b',          r'\bcryptic splicing\b',
            r'\bsplicing defect\b',            r'\bsplicing error\b',
            r'\bpseudoexon\b',                 r'\bpoison exon\b',
        ],
        '剪接靶向治疗': [
            r'\bsplicing modulator\b',
            r'\bsplicing.{0,30}(?:drug|therap\w+|inhibitor|compound)\b',
            r'\bantisense oligonucleotide\b',
            r'\bnusinersen\b', r'\brisdiplam\b', r'\bspliceostatin\b',
            r'\bsplice.correcting\b', r'\bsplice.switching\b',
        ],
        '剪接调控研究': [
            r'\bsplicing factor\b',            r'\bsplicing regul\w+\b',
            r'\bexonic splicing enhancer\b',   r'\bexonic splicing silencer\b',
            r'\bintronic splicing\b',
            r'\b(?<!\w)ESE(?!\w)\b',           r'\b(?<!\w)ESS(?!\w)\b',
            r'\bpolypyrimidine tract\b',
            r'\bSRSF\b', r'\bhnRNP\b', r'\bSR protein\b',
        ],
        '剪接生物信息学/工具': [
            r'\bsplicing.{0,20}predict\w+\b', r'\bpredict\w+.{0,20}splicing\b',
            r'\bsplicing.{0,20}(?:algorithm|tool|model|software)\b',
            r'\bspliceAI\b', r'\bdifferential splicing\b',
            r'\bsplicing quantitative\b', r'\bsQTL\b',
            r'\bRNA.seq.{0,30}splicing\b',
        ],
    },

    'SPECIES_PATTERNS': {
        '人类/临床': [r'\bhuman[s]?\b', r'\bpatient[s]?\b', r'\bclinical\b',
                      r'\bHomo sapiens\b', r'\bhuman cell\b'],
        '小鼠/大鼠': [r'\bmouse\b', r'\bmice\b', r'\bmurine\b', r'\brat\b',
                      r'\bMus musculus\b', r'\bRattus\b'],
        '酵母':      [r'\byeast\b', r'\bSaccharomyces\b', r'\bS\. cerevisiae\b'],
        '植物':      [r'\bplant\b', r'\bArabidopsis\b', r'\bmaize\b', r'\brice\b'],
        '斑马鱼':    [r'\bzebrafish\b', r'\bDanio rerio\b'],
        '果蝇':      [r'\bDrosophila\b', r'\bfruit fly\b'],
        '线虫':      [r'\bC\. elegans\b', r'\bnematode\b'],
        '细胞系':    [r'\bcell line[s]?\b', r'\bin vitro\b', r'\bcultured cell\b'],
        '计算/综述': [r'\bin silico\b', r'\bcomputational\b',
                      r'\breview\b', r'\bmeta.analysis\b'],
    },
}


# ══════════════════════════════════════════════════════════════════════
# 领域配置：虚拟细胞（新增）
# 涵盖：全细胞模型、代谢建模(FBA/GEM)、信号转导模型、随机/空间模拟、
#        细胞力学、多细胞/组织模型、基因调控网络、AI细胞基础模型
# ══════════════════════════════════════════════════════════════════════
VCELL = {
    'name':           '虚拟细胞',
    'sheet_title':    '虚拟细胞分析',
    'conclusion_col': '虚拟细胞相关结论',
    'fallback_label': '虚拟细胞相关（其他）',
    'header_color':   '1D6B3E',   # 深绿，象征细胞/生命
    'stat_color':     '1D6B3E',
    'stripe_color':   'D5E8D4',
    'out_path':       'E:/Project/BioInfo/temp/sets_VirtualCell.xlsx',
    'col_widths':     [55, 75, 15, 16, 10, 16, 32, 22, 88],

    'TIER1': [
        # ── 直接命名的虚拟细胞工具 / 概念 ────────────────────────────
        (r'\bvirtual cell\b',                        'virtual cell'),
        (r'\bVCell\b',                               'VCell(software)'),
        (r'\bwhole.?cell model[s]?\b',               'whole-cell model'),
        (r'\bwhole.?cell simulation[s]?\b',          'whole-cell simulation'),
        (r'\bwhole.?cell modelling\b',               'whole-cell modelling'),
        (r'\bdigital cell[s]?\b',                    'digital cell'),
        (r'\bdigital twin.{0,20}cell\b',             'digital twin of cell'),
        (r'\bcell.{0,15}digital twin\b',             'cell digital twin'),
        (r'\bin silico cell\b',                      'in silico cell'),
        (r'\bcomputational cell model\b',            'computational cell model'),
        (r'\bcell.level simulation\b',               'cell-level simulation'),
        (r'\bsynthetic cell model\b',                'synthetic cell model'),
        (r'\bminimal cell model\b',                  'minimal cell model'),
        (r'\bcomputational model.{0,20}(?:the )?cell\b', 'computational model of cell'),
        (r'\bcell simulation[s]?\b',                 'cell simulation'),
        # ── 主流模拟软件（本身即代表"虚拟细胞"实践）─────────────────
        (r'\bE.Cell\b',                              'E-Cell(software)'),
        (r'\bEcell\b',                               'Ecell(software)'),
        (r'\bMCell\b',                               'MCell(Monte Carlo cell)'),
        (r'\bSmoldyn\b',                             'Smoldyn'),
        (r'\bSpringsal[Aa][Dd]\b',                   'SpringSaLaD'),
        (r'\bPhysiCell\b',                           'PhysiCell'),
        (r'\bCompuCell3[Dd]\b',                      'CompuCell3D'),
        (r'\b(?<!\w)COPASI\b',                       'COPASI'),
        (r'\bBioNetGen\b',                           'BioNetGen'),
        (r'\bChaste\b',                              'Chaste(cell simulation)'),
    ],

    'TIER2': [
        # ── 代谢建模 ─────────────────────────────────────────────────
        (r'\bflux balance analysis\b',               'flux balance analysis(FBA)'),
        (r'\b(?<!\w)FBA\b',                          'FBA(flux balance analysis)'),
        (r'\bgenome.scale metabolic model\b',        'genome-scale metabolic model(GEM)'),
        (r'\b(?<!\w)GEM(?!\w)\b',                    'GEM(genome-scale metabolic model)'),
        (r'\bGSMM\b',                                'GSMM'),
        (r'\bconstraint.based.{0,20}(?:model|reconstruct|analys)\w*\b',
                                                     'constraint-based model'),
        (r'\b(?<!\w)COBRA\b',                        'COBRA toolbox'),
        (r'\bmetabolic flux\b',                      'metabolic flux'),
        (r'\bflux variability analysis\b',           'flux variability analysis(FVA)'),
        (r'\b(?<!\w)FVA\b',                          'FVA'),
        (r'\bdynamic FBA\b',                         'dynamic FBA(dFBA)'),
        (r'\bparsimonious FBA\b',                    'parsimonious FBA'),
        # ── 信号转导模型 ──────────────────────────────────────────────
        (r'\bcell signaling.{0,20}model\b',          'cell signaling model'),
        (r'\bsignaling network model\b',             'signaling network model'),
        (r'\bkinetic model.{0,30}(?:signal|cell|receptor)\b',
                                                     'kinetic signaling model'),
        (r'\bcell cycle model\b',                    'cell cycle model'),
        (r'\breceptor.ligand.{0,20}model\b',         'receptor-ligand model'),
        (r'\bsecond messenger model\b',              'second messenger model'),
        # ── 随机 / 空间模型 ───────────────────────────────────────────
        (r'\bstochastic simulation algorithm\b',     'stochastic simulation algorithm(SSA)'),
        (r'\b(?<!\w)SSA\b',                          'SSA(stochastic simulation)'),
        (r'\bGillespie algorithm\b',                 'Gillespie algorithm'),
        (r'\breaction.diffusion model\b',            'reaction-diffusion model'),
        (r'\bspatial cell model\b',                  'spatial cell model'),
        (r'\bspatial stochastic\b',                  'spatial stochastic simulation'),
        # ── 基因调控网络模型 ─────────────────────────────────────────
        (r'\bgene regulatory network.{0,20}model\b', 'gene regulatory network model'),
        (r'\bBoolean network model\b',               'Boolean network model'),
        (r'\b(?<!\w)GRN model\b',                    'GRN model'),
        # ── 细胞力学 ─────────────────────────────────────────────────
        (r'\bcell mechanics model\b',                'cell mechanics model'),
        (r'\bcytoskeletal model\b',                  'cytoskeletal model'),
        (r'\bcell membrane model\b',                 'cell membrane model'),
        (r'\bbiomechanical model.{0,20}cell\b',      'biomechanical cell model'),
        (r'\bcell migration model\b',                'cell migration model'),
        # ── 标准格式 / 框架 ───────────────────────────────────────────
        (r'\bSystems Biology Markup Language\b',     'SBML'),
        (r'\b(?<!\w)SBML\b',                         'SBML'),
        (r'\bCellML\b',                              'CellML'),
        (r'\b(?<!\w)BNGL\b',                         'BNGL(BioNetGen language)'),
        (r'\b(?<!\w)SBGN\b',                         'SBGN'),
        # ── AI 细胞基础模型 ───────────────────────────────────────────
        (r'\bsingle.cell foundation model\b',        'single-cell foundation model'),
        (r'\bcell foundation model\b',               'cell foundation model'),
        (r'\bscGPT\b',                               'scGPT'),
        (r'\bGeneformer\b',                          'Geneformer'),
        (r'\bUniversal Cell Embedding\b',            'Universal Cell Embedding(UCE)'),
        (r'\b(?<!\w)UCE\b',                          'UCE(Universal Cell Embedding)'),
        (r'\bvirtual patient\b',                     'virtual patient'),
    ],

    'TIER2_GATED': [
        # ── 系统生物学（泛指，需细胞建模语境确认）────────────────────
        (r'\bsystems biology\b',                     'systems biology'),
        # ── 数学模型（通用词，需限定在细胞语境）──────────────────────
        (r'\bmathematical model.{0,20}cell\b',       'mathematical model of cell'),
        (r'\bODE model\b',                           'ODE model'),
        (r'\bordinary differential equation.{0,20}(?:model|cell|signal)\b',
                                                     'ODE cell model'),
        (r'\bstochastic model.{0,20}cell\b',         'stochastic cell model'),
        (r'\bcomputational model.{0,15}(?:biolog|cell)\w*\b',
                                                     'computational biology model'),
        # ── 基因调控（无 "model" 限定时歧义大）────────────────────────
        (r'\bgene regulatory network\b',             'gene regulatory network(GRN)'),
        (r'\b(?<!\w)GRN\b',                          'GRN'),
        (r'\bBoolean network\b',                     'Boolean network'),
        (r'\bPetri net.{0,20}(?:biolog|cell|model)\b', 'biological Petri net'),
        # ── 细胞自动机 ────────────────────────────────────────────────
        (r'\bcellular automaton\b',                  'cellular automaton'),
        # ── 多细胞 / 器官（宽泛，需语境确认为计算模型）────────────────
        (r'\bmulticellular model\b',                 'multicellular model'),
        (r'\btissue model\b',                        'tissue model'),
        (r'\borgan model\b',                         'organ model'),
        # ── 粗粒化 / 分子动力学（限细胞/膜语境）──────────────────────
        (r'\bmolecular dynamics.{0,20}(?:cell|membran|protein)\b',
                                                     'molecular dynamics(cell)'),
        (r'\bcoarse.grained model.{0,20}(?:cell|membran)\b',
                                                     'coarse-grained cell model'),
        # ── 机制模型（宽泛）─────────────────────────────────────────
        (r'\bmechanistic model.{0,20}(?:cell|signal|metabol)\b',
                                                     'mechanistic model(cell)'),
    ],

    'CONTEXT': [
        r'\bvirtual cell\b',
        r'\bVCell\b',
        r'\bwhole.?cell model\b',
        r'\bcell simulation\b',
        r'\bflux balance\b',
        r'\bGillespie\b',
        r'\b(?<!\w)SBML\b',
        r'\bCellML\b',
        r'\bCompuCell\b',
        r'\bPhysiCell\b',
        r'\b(?<!\w)COPASI\b',
        r'\bBioNetGen\b',
        r'\bcell signaling.{0,20}model\b',
        r'\bgenome.scale metabolic\b',
        r'\bcomputational cell\b',
        r'\bcell.level model\b',
        r'\bin silico cell\b',
        r'\bdigital.{0,10}cell\b',
        r'\bscGPT\b',
        r'\bGeneformer\b',
        r'\bsingle.cell foundation\b',
        r'\bE.Cell\b',
        r'\bSmoldyn\b',
        r'\bSpringsal\w+\b',
        r'\bMCell\b',
    ],

    'TIER3': [
        (r'\bcell model[s]?\b',          'cell model'),
        (r'\bcomputational\b',           'computational'),
        (r'\bsimulation[s]?\b',          'simulation'),
        (r'\bmathematical model\b',      'mathematical model'),
        (r'\bnetwork model\b',           'network model'),
        (r'\bsignaling pathway\b',       'signaling pathway'),
        (r'\bmetabolic model\b',         'metabolic model'),
        (r'\bbiological network\b',      'biological network'),
        (r'\bsystems biology\b',         'systems biology'),
        (r'\bpathway model\b',           'pathway model'),
    ],

    'FALSE_POSITIVE_PATTERNS': [
        (r'\bclinical trial simulation\b',    'clinical trial simulation（临床模拟）'),
        (r'\bdrug pharmacokinetic[s]?\b',     'drug pharmacokinetics（药代动力学）'),
        (r'\bweather model\b',               'weather model（气象模型）'),
        (r'\beconomic model\b',              'economic model（经济模型）'),
        (r'\btraffic simulation\b',          'traffic simulation（交通模拟）'),
        (r'\bepidemiological model\b',       'epidemiological model（流行病学模型）'),
        (r'\bpopulation model\b',            'population model（种群模型）'),
        (r'\bclimate model\b',               'climate model（气候模型）'),
        (r'\bsocial network\b',              'social network（社交网络）'),
        (r'\bfinancial model\b',             'financial model（金融模型）'),
        (r'\bdeep learning\b',               'deep learning（泛化ML，非细胞专用）'),
        (r'\bneural network[s]?\b',          'neural network（人工神经网络）'),
    ],

    'CATEGORY_PATTERNS': {
        '代谢与通量建模': [
            r'\bflux balance analysis\b', r'\b(?<!\w)FBA\b',
            r'\bgenome.scale metabolic\b', r'\b(?<!\w)GEM\b',
            r'\b(?<!\w)COBRA\b', r'\bmetabolic flux\b',
            r'\bflux variability\b', r'\bconstraint.based\b',
        ],
        '信号转导网络模型': [
            r'\bcell signaling.{0,20}model\b', r'\bsignaling network model\b',
            r'\bkinetic model\b', r'\bcell cycle model\b',
            r'\breceptor.{0,15}model\b', r'\bODE model\b',
            r'\bsecond messenger\b',
        ],
        '基因调控网络模型': [
            r'\bgene regulatory network\b', r'\bBoolean network\b',
            r'\b(?<!\w)GRN\b', r'\btranscription.{0,20}model\b',
            r'\bregulatory network.{0,20}(?:model|simulat)\b',
            r'\bGeneformer\b', r'\bscGPT\b',
        ],
        '全细胞与最小细胞模型': [
            r'\bwhole.?cell model\b', r'\bminimal cell\b',
            r'\bVCell\b', r'\bE.Cell\b', r'\bEcell\b',
            r'\bwhole.?cell simulation\b', r'\bdigital cell\b',
        ],
        '细胞力学与形态模型': [
            r'\bcell mechanics\b', r'\bcytoskeletal\b',
            r'\bcell membrane model\b', r'\bbiomechanical model\b',
            r'\bCompuCell3[Dd]\b', r'\bcell morpholog\w+\b',
            r'\bcell shape\b', r'\bcell migration model\b',
        ],
        '多细胞与组织模型': [
            r'\bmulticellular model\b', r'\btissue model\b',
            r'\bPhysiCell\b', r'\bcell population model\b',
            r'\btumor.{0,20}model\b', r'\borgan.on.chip model\b',
            r'\borganoid.{0,20}model\b',
        ],
        '随机与空间模型': [
            r'\bstochastic simulation\b', r'\bGillespie\b',
            r'\b(?<!\w)SSA\b', r'\breaction.diffusion\b',
            r'\bspatial stochastic\b', r'\bMCell\b',
            r'\bSmoldyn\b', r'\bSpringsal\w+\b',
        ],
        'AI细胞基础模型': [
            r'\bscGPT\b', r'\bGeneformer\b',
            r'\bfoundation model.{0,20}(?:cell|single.cell)\b',
            r'\bsingle.cell.{0,20}foundation model\b',
            r'\bUniversal Cell Embedding\b', r'\b(?<!\w)UCE\b',
            r'\bpre.trained.{0,20}(?:cell|single.cell)\b',
        ],
    },

    'SPECIES_PATTERNS': {
        '人类/临床': [r'\bhuman[s]?\b', r'\bpatient[s]?\b', r'\bclinical\b',
                      r'\bHomo sapiens\b', r'\bhuman cell\b'],
        '小鼠/大鼠': [r'\bmouse\b', r'\bmice\b', r'\bmurine\b', r'\brat\b',
                      r'\bMus musculus\b'],
        '酵母':          [r'\byeast\b', r'\bSaccharomyces\b', r'\bS\. cerevisiae\b'],
        '大肠杆菌/原核': [r'\bE\. coli\b', r'\bEscherichia\b', r'\bprokaryot\w+\b',
                          r'\bBacillus\b'],
        '斑马鱼':    [r'\bzebrafish\b', r'\bDanio rerio\b'],
        '果蝇':      [r'\bDrosophila\b', r'\bfruit fly\b'],
        '线虫':      [r'\bC\. elegans\b', r'\bnematode\b'],
        '细胞系/体外': [r'\bcell line[s]?\b', r'\bin vitro\b', r'\bcultured cell\b',
                        r'\bHeLa\b', r'\bHEK293\b'],
        '计算/综述': [r'\bin silico\b', r'\bcomputational\b',
                      r'\breview\b', r'\bmeta.analysis\b'],
    },
}


# ══════════════════════════════════════════════════════════════════════
# 领域注册表
# ══════════════════════════════════════════════════════════════════════
DOMAINS = {
    'copd':         COPD,
    'neuro':        NEURO,
    'rna-splicing': RNA,
    'vcell':        VCELL,
}


# ══════════════════════════════════════════════════════════════════════
# 入口
# ══════════════════════════════════════════════════════════════════════
if __name__ == '__main__':
    keys = sys.argv[1:]
    if not keys:
        keys = list(DOMAINS.keys())   # 默认全部运行

    unknown = [k for k in keys if k not in DOMAINS]
    if unknown:
        print('未知领域标识: {}，可用: {}'.format(', '.join(unknown), ', '.join(DOMAINS)))
        sys.exit(1)

    for key in keys:
        run_domain(DOMAINS[key], SRC_PATH)
